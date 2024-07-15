"""
This code is used to assign tasks to 5 people. The tasks include the results of 2 large models and the results of humans.
The output is 5 xlsx files, each file contains the results of 2 large models and humans.
The header is: id, model1, comprehensiveness, hallucination, accuracy, expressiveness, accept without revision, model2, comprehensiveness, hallucination, accuracy, expressiveness, accept without revision, human, comprehensiveness, hallucination, accuracy, expressiveness, accept without revision
"""

import pandas as pd
import numpy as np
import string
import random
import re
import os
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# File path
model1_output_path = "../data/response_wxyy20240620.csv"
model2_output_path = "../data/response_claude35sonnet20240620.csv"
outfile = "./data/task.xlsx"

# Read data
model1_output = pd.read_csv(model1_output_path, encoding="gbk")
model2_output = pd.read_csv(model2_output_path, encoding="gbk")
model3_output = model1_output.copy()  # Model1 and Model2 both contain the human's impression, so choose one of them

# 将index设置为检查号
model1_output.set_index("检查号", inplace=True)
model2_output.set_index("检查号", inplace=True)
model3_output.set_index("检查号", inplace=True)

# 去掉index重复的行
model1_output = model1_output[~model1_output.index.duplicated(keep='first')]
model2_output = model2_output[~model2_output.index.duplicated(keep='first')]
model3_output = model3_output[~model3_output.index.duplicated(keep='first')]

# 查看哪些index在两个模型的输出中都存在
common_index = model1_output.index.intersection(model2_output.index)
model1_output = model1_output.loc[common_index]

# 给每个人每个模型的输出分配一个随机的id：比如第一个人的第一个模型的输出的id为s1_m876，其中s1表示第一个人，m876表示第一个模型的输出，但876是随机的
task = pd.DataFrame(columns=["影像所见", "model1", "完整性1", "幻觉性1", "准确性1", "表达力1", "修改度1", "model2", "完整性2", "幻觉性2", "准确性2", "表达力2", "修改度2", "model3", "完整性3", "幻觉性3", "准确性3", "表达力3", "修改度3"], index=model1_output.index)
for i, idx in enumerate(task.index):
    task.loc[idx, "影像所见"] = model1_output.loc[idx, "影像所见"]
    
    random_id = f"s{idx}" + "_m" + ''.join(random.choice(string.ascii_letters) for _ in range(6))
    impression = f"${random_id}$" + "\n" + model1_output.loc[idx, "AI生成的印象"]
    task.loc[idx, "model1"] = impression

    random_id = f"s{idx}" + "_m" + ''.join(random.choice(string.ascii_letters) for _ in range(6))
    impression = f"${random_id}$" + "\n" + model2_output.loc[idx, "AI生成的印象"]
    task.loc[idx, "model2"] = impression

    random_id = f"s{idx}" + "_m" + ''.join(random.choice(string.ascii_letters) for _ in range(6))
    impression = f"${random_id}$" + "\n" + model3_output.loc[idx, "印象"]
    task.loc[idx, "model3"] = impression

# 生成一个对应随机id的dict，以返回到原始的检查号
match_id = {}
for i, idx in enumerate(task.index):
    match_id[idx] = {}
    match_id[idx]['wxyy'] = task.loc[idx, 'model1'].split("$")[1]
    match_id[idx]['claude'] = task.loc[idx, 'model2'].split("$")[1]

# 保存match_id 换行
if not os.path.exists("./data"):
    os.makedirs("./data")
with open("./data/match_id.json", "w") as f:
    json.dump(match_id, f, indent=4)

# 将task中的每一行model1、model2、model3列的顺序打乱
for i, idx in enumerate(task.index):
    l = [task.loc[idx, "model1"], task.loc[idx, "model2"], task.loc[idx, "model3"]]
    random.shuffle(l)
    task.loc[idx, "model1"] = l[0]
    task.loc[idx, "model2"] = l[1]
    task.loc[idx, "model3"] = l[2]

# 随机取出50份报告做观察者间一致性分析
task_for_icc =task.sample(50)
# save icc index
with open("./data/icc_index.json", "w") as f:
    json.dump(task_for_icc.index.tolist(), f, indent=4)

# Delete the 50 reports from the original task
task.drop(task_for_icc.index, inplace=True)

# 把task拆分成5行数一样的小表格
task1 = task.iloc[:int(len(task)/5)]
task2 = task.iloc[int(len(task)/5):int(len(task)/5)*2]
task3 = task.iloc[int(len(task)/5)*2:int(len(task)/5)*3]
task4 = task.iloc[int(len(task)/5)*3:int(len(task)/5)*4]
task5 = task.iloc[int(len(task)/5)*4:]

# Concatenate the 50 reports to the 5 tasks
task1 = pd.concat([task1, task_for_icc])
task2 = pd.concat([task2, task_for_icc])
task3 = pd.concat([task3, task_for_icc])
task4 = pd.concat([task4, task_for_icc])
task5 = pd.concat([task5, task_for_icc])


def save(outfile, task):
    # 冻结表头和index
    task.to_excel(outfile, index=True, freeze_panes=(1, 1))
    # 打开工作簿以进行进一步的格式设置
    book = load_workbook(outfile)
    sheet = book.active

    # 定义最大列宽
    max_column_width = 15

    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            
            # 1. 自动换行
            cell.alignment = Alignment(wrap_text=True)
        
        # 3. 调整列宽，但设置最大宽度
        adjusted_width = min((max_length + 2), max_column_width)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # 保存更改
    book.save(outfile)

excutors = ["陈秀珍", "段亚妮", "陈耀萍", "黎超", "董梦实"]
for i, task in enumerate([task1, task2, task3, task4, task5]):
    save(outfile.replace(".xlsx", f"_{excutors[i]}.xlsx"), task)
